#!/usr/bin/python3
#
#  Author:  Steve Mann
#  Date:    05-02-2017
#  Rev:     -
#
#  Description:
#  Scan through "Timing Patterns" tab of the spreadsheet entered in the command line.  Do the following:
#   1)  Transpose the top labels and data so that the labels are on the left (from the top)
#        and the data goes to the right with time (basically rotate CCW 90 degrees).  Enter this into
#        the "Transpose Timing Patterns" tab.
#   2)  Copy the transposed data to a third tab "Transpose Timing Plot"  Do the following to the data:
#           A) Add a row between each data set.
#           B) Set a thick border above all 1s, a thick border below all 0s, and a thick border between
#               any transitions from 0 to 1 or 1 to 0.
#           C) Hide (or very light shade of text) the 1s and 0s.
#           D) Highlight with light shade every Xth row for ease of reading (visual)
#           E) Highlight with light shade every Xth column for ease of reading (visual)
#           F) Option to hide all rows that do not have labels (unused bits)
#
#  Execution:
#  Command line examples:
#
#    To clear the previous plot and plot new:
#    >python TranslateH4RGClocking.py H4RGb.xlsx
#
#    To **only** clear the old plot:
#    >python TranslateH4RGClocking.py H4RGb.xlsx ClearOnly
#
#
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Color, PatternFill, Font
import os
import sys
import re
import time
import datetime
import argparse

print("\n******* Start H4RG Timing Plot Generation  *******\n")

# Define where the spreadsheet values are kept
BeginInputRow    =  8                    # This will never change - rows will always start in row 8
BeginInputRow_Data = BeginInputRow + 1
BeginInputColumn =  5                    # This will never change - there will always be 48 control signals starting in column 5 (E)
EndInputColumn   = 53                    # This will never change - there will always be 48 control signals ending in column 53 (BA)
EndInputColumnActual = EndInputColumn - 1     # EndInputColumn is the upper bound **that does not get met** in the list.  The actual last item is EndInputColumn - 1
StateChangeColumn = 3
HexColumn = 4                          # Row of the input sheet where the Hex bit pattern representations are saved.
BeginOutputRow = 5
BeginOutputColumn = 1
BeginOutputPlotRow = 5
BeginOutputPlotColumn = 3
OutputPlotTimingLabelRow = 3

# Border shading...
#   Only one shading set is allowed per cell. i.e. writing a top border then bottom border on the same cell will result
#    in **only** the bottom border being highlited.
#   Therefore... step across cells, highliting the top/bottom **and** left border (based on cell to the left and current
#    cell) at the same time.  On first data cell of row, just highlight top/bottom since there is no left cell value.
#
thick_top_border = Border(top=Side(style='thick'))
thick_bottom_border = Border(bottom=Side(style='thick'))
thick_top_left_border = Border(top=Side(style='thick'),
                               left=Side(style='thick'))
thick_bottom_left_border = Border(bottom=Side(style='thick'),
                                  left=Side(style='thick'))
no_borders = Border(top=Side(style=None),
                    left=Side(style=None),
                    bottom=Side(style=None),
                    right=Side(style=None))
thick_left_thin_bottom_border = Border(left=Side(style='thick'),
                                       bottom=Side(style='thin'))
thin_bottom_border = Border(bottom=Side(style='thin'))

# Set up all fill colors
PlotTimeMarkerFill = PatternFill("solid", fgColor="f2f287")
NoCellColorFill = PatternFill("none")

# Set up light shade font for plot 1s and 0s
FontLight = Font(color='d1d1d1')      # 0xd1d1d1 = Very light grey
FontBlackBold = Font(bold=True, color='000000')  # 0xd1d1d1 = Black,

# Create a timestamp for future use and print it
now = datetime.datetime.now()
print("Date and Time =", now)

# Capture command line arguments
parser = argparse.ArgumentParser()
parser.add_argument("file1_name", nargs='?', default="empty_string")
parser.add_argument("clear_only", nargs='?', default="empty_string")
args = parser.parse_args()
print(args)
print(parser)
print(file1_name)

# Set clear_only to all lower case for testing against
clear_only = args.clear_only.lower()


# If no file name to process, display error and get out.
if (args.file1_name == 'empty_string'):
    print("\nERROR: No parameters given! Please add file name to process.")
    sys.exit()
# If the file is open, tell user to close file and try again.    
elif args.file1_name.closed() is False:
    print("\nERROR: The file is open! Please close it and run again.")
    sys.exit()      
else:
    # Save off file name arguments
    file1_name = args.file1_name
    print("\nWorking File Name = %s" % file1_name)

    if (args.clear_only == 'empty_string'):
        print("\nClear then process\n")
    else:
        if( clear_only != "clearonly" ):
            print("\nERROR - Invalid Clear (2nd) parameter: \"%s\".  Must be any case of \"ClearOnly\" to clear only.  No 2nd parameter = clear, then process." % args.clear_only)
            sys.exit()
        else:
            print("\n******  Clear ONLY  ******\n")

    wb = openpyxl.load_workbook(file1_name)

    sheet_input = wb.get_sheet_by_name("Timing Patterns")
    print("Working Input Sheet Title =", sheet_input.title)
    sheet_output = wb.get_sheet_by_name("Transpose Timing Patterns")
    print("Working Output Timing Pattern Sheet Title =", sheet_output.title)
    sheet_output_plot = wb.get_sheet_by_name("Transpose Timing Plot")
    print("Working Output Timing Plot Sheet Title =", sheet_output_plot.title)

    #
    #  Spreadsheet extractable data.  Row 8 contains the labels.  Row 9 - XXXX contains the clocks
    #  To transpose - put the label for bit 47 (Row 8, Col 5) in the 47th row down (plus offset)
    #
    # Note regarding the range function: range(2,6) will return 2, 3, 4, 5. It does not return the value 6.
    #    range returns a sequence up to but NOT including the upper bound.


    #
    # Count the number of input rows with data in them.  Use the Bit 0 column since it will always be used for control (I think)
    #
    i_row = BeginInputRow_Data               # The first data row (row 9)
    EndRow = BeginInputRow_Data              # This is the actual row, not a row count
    while (sheet_input.cell(row=i_row,column=EndInputColumnActual).value != None):
        i_row = i_row + 1
        EndRow = EndRow + 1

    EndRowActual = EndRow - 1           # EndRow is the upper bound **that does not get met** in the list.  The actual last item is EndRow - 1

    #
    # Clear out the two sheets to be written as well as the Hex column in the input sheet.
    #   Clear the timing plots (borders) as well.
    print("\nClearing output sheets...\n")
    for j_column in range(1,EndInputColumn+50):             # Clear output sheets from column 1 to EndInputColumn+50
        for i_row in range(1,EndRow+50):               # Clear output sheets from row 1    to EndRow+50
            # Clear the output plot sheet (data and borders)
            sheet_output_plot.cell(row=i_row,column=j_column).value = None
            sheet_output_plot.cell(row=i_row,column=j_column).border = no_borders
            sheet_output_plot.cell(row=i_row,column=j_column).fill = NoCellColorFill
            #Clear the transposed, full data, sheet
            sheet_output.cell(row=i_row,column=j_column).value = None
            #Clear the Hex and changed state columns of the input sheet
            if (j_column == StateChangeColumn) and (i_row > 8):
                sheet_input.cell(row=i_row,column=StateChangeColumn).value = None
            if (j_column == HexColumn) and (i_row > 8):
                sheet_input.cell(row=i_row,column=HexColumn).value = None
    #
    # If we're only supposed to clear, print that we're done, save the file, then get out here.
    #
    if( clear_only == "clearonly" ):
        print("\n******* Completed H4RG Timing Plot CLEAR ONLY  *******\n")
        wb.save(file1_name)
        sys.exit()

    # Count the number of labels so that we know where to start the plots.  Put LSB at the top, MSB at the bottom.
    # Only plot the rows with labels.
    Labels = 0
    for j_column in range(BeginInputColumn,EndInputColumn):                                   # Start with column 5 (E).  Run down column, checking rows. Repeat through column 52 (AZ)
        if (sheet_input.cell(row=BeginInputRow,column=j_column).value) != None:
            Labels = Labels + 1

            # Start the plots at the "bottom" of the page with the MSB, working our way up every other row to the LSB.
    PlotRow = (Labels * 2) + (BeginOutputRow - 2)
    PlotRowBottom = PlotRow

    #
    # Transpose, create two plots sheets - one with just labeled rows, and plot the only labeled row sheet.
    #
    for j_column in range(BeginInputColumn,EndInputColumn):           # Start with column 5 (E).  Run down column, checking rows. Repeat through column 52 (AZ)
        for i_row in range(BeginInputRow,EndRow):                # Start at row 8, run through 28

            # Build "Transpose Timing Patterns" sheet

            # While we're running through all of the data, verify that data is either 1 or 0.  If not, print error message, save fiel, then exit.
            if ( i_row != BeginInputRow ):   # Don't check labels to verify whether they are 1 or 0
                if ( (sheet_input.cell(row=i_row,column=j_column).value != 0) and (sheet_input.cell(row=i_row,column=j_column).value != 1) ):
                    print("\nERROR - Invalid bit patern data (not a 1 or 0) in cell row:", i_row,", column:", j_column,". Value = ", sheet_input.cell(row=i_row,column=j_column).value)
                    wb.save(file1_name)
                    sys.exit()

            sheet_output.cell(row=(EndInputColumn+HexColumn-j_column),column=(i_row-6)).value = sheet_input.cell(row=i_row,column=j_column).value


            # Build "Transpose Timing Plot" sheet
            # If the first value in the column is a label (not None) then build the row.  Skip one row between each row built for plotting.
            if (sheet_input.cell(row=BeginInputRow,column=j_column).value) != None:

                # Set the output data font to light grey, labels to black, bold.
                if (i_row == BeginInputRow):
                    sheet_output_plot.cell(row=(PlotRow),column=(i_row-6)).font = FontBlackBold
                else:  # else we're dealing with output data (should be 1s or 0s)
                    sheet_output_plot.cell(row=(PlotRow),column=(i_row-6)).font = FontLight

                sheet_output_plot.cell(row=(PlotRow),column=(i_row-6)).value = sheet_input.cell(row=i_row,column=j_column).value


                # While populating rows (plot) add thick border above 1s, below 0s, and between 1 to 0 and 0 to 1 transitions
                #   thick_top_border
                #   thick_bottom_border
                #   thick_top_left_border
                #   thick_bottom_left_border
                # Also, center data values.
                # And... make data font a light shade
                #
                if i_row == BeginInputRow:               # Don't do anything.  This is the label
                    sheet_output_plot.cell(row=(PlotRow),column=(i_row-6)).alignment = Alignment(horizontal="left")   # Align cell (the label) left
                    print("Plotting...")
                elif i_row == BeginInputRow + 1:   # Don't do left border. This is the first data cell.
                    sheet_output_plot.cell(row=(PlotRow),column=(i_row-6)).alignment = Alignment(horizontal="center") # Align cell (data) center
                    if sheet_output_plot.cell(row=(PlotRow),column=(i_row-6)).value == 0:
                        sheet_output_plot.cell(row=(PlotRow),column=(i_row-6)).border = thick_bottom_border
                    elif  sheet_output_plot.cell(row=(PlotRow),column=(i_row-6)).value == 1:
                        sheet_output_plot.cell(row=(PlotRow),column=(i_row-6)).border = thick_top_border
                    else:
                        print("Plot error: Value not 0 or 1 in Transpose Timing Plot cell, column:", i_row-6, " row:", PlotRow)
                else:                           # Do left border
                    sheet_output_plot.cell(row=(PlotRow),column=(i_row-6)).alignment = Alignment(horizontal="center") # Align cell (data) center
                    if sheet_output_plot.cell(row=(PlotRow),column=(i_row-6)).value == 0:
                        if sheet_output_plot.cell(row=(PlotRow),column=(i_row-7)).value == 1:
                            sheet_output_plot.cell(row=(PlotRow),column=(i_row-6)).border = thick_bottom_left_border
                        else:
                            sheet_output_plot.cell(row=(PlotRow),column=(i_row-6)).border = thick_bottom_border
                    elif  sheet_output_plot.cell(row=(PlotRow),column=(i_row-6)).value == 1:
                        if sheet_output_plot.cell(row=(PlotRow),column=(i_row-7)).value == 0:
                            sheet_output_plot.cell(row=(PlotRow),column=(i_row-6)).border = thick_top_left_border
                        else:
                            sheet_output_plot.cell(row=(PlotRow),column=(i_row-6)).border = thick_top_border
                    else:
                        print("Plot error on Transpose Timing Plot cell", i_row-6, PlotRow)

                if i_row == EndRowActual:               # The row has completed so move up two to the next row
                    PlotRow = PlotRow - 2




    #
    # Add bit numbers to column A of "Transpose Timing Patterns" sheet
    #
    sheet_output.cell(row=BeginOutputRow-1,column=BeginOutputColumn).value = "Bit Number"       # Go up one (BeginOutputRow-1) for the label
    for j_column in range(BeginInputColumn,EndInputColumn):       # Start with column 5 (E), run through 53 (BA)
        sheet_output.cell(row=j_column,column=BeginOutputColumn).value = j_column-BeginOutputRow      # 47 bits starting at 0

    #
    # Add timing information to OutputPlotTimingLabelRow of sheet_output_plot.  Start with 0 at C3 (3,3), continuing to the end of the columns
    #   Only add for 0, 5, 10, 15, etc...  (1st and every 10th).  Use i for this.
    #   Write "Time (us)" label to B,3
    #   EndRow is the calculated length of the data set
    #   Use thick_left_thin_bottom_border for cells with numbers
    #   Use thin_bottom_border for cells without numbers
    #   Shade with light color the column that has the timing number in it
    sheet_output_plot.cell(row=OutputPlotTimingLabelRow,column=(BeginOutputPlotColumn-1)).value = "Time (us)"
    sheet_output_plot.cell(row=OutputPlotTimingLabelRow,column=BeginOutputPlotColumn-1).border = thin_bottom_border
    i = 0
    for j_column in range(BeginOutputPlotColumn,(BeginOutputPlotColumn+EndRow)):
        sheet_output_plot.cell(row=OutputPlotTimingLabelRow,column=j_column).border = thin_bottom_border
        if( i % 10 == 0 ):
            sheet_output_plot.cell(row=OutputPlotTimingLabelRow,column=j_column).border = thick_left_thin_bottom_border   # Overwrites the thin bottom border only
            sheet_output_plot.cell(row=OutputPlotTimingLabelRow,column=j_column).value = i/2
            for j in range(OutputPlotTimingLabelRow,PlotRowBottom+1):
                sheet_output_plot.cell(row=j,column=j_column).fill = PlotTimeMarkerFill
        i = i + 1

    # Convert 48 bit binary across each row to hex.  48 bit binary = 12 hex digits (0xFFFFFFFFFFFF)
    #  Clear the Result.  Go through each sheet_input row from left to right, adding cell value ^^2 to the Result.  Convert the result to Hex.
    # BeginInputColumn =  5
    # EndInputColumn   = 53 (EndInputColumnActual = 52)
    #
    #   5          6          7          8          9          10       ....    51          52
    # 2**(52-5)  2**(52-6)  2**(52-7)  2**(52-8)  2**(52-9)  2**(52-10)       2**(52-51)  2**(52-52)
    #
    # For each column in the row, if value == 1:
    #    FortyEightBitResult = FortyEightBitResult + 2**(EndInputColumnActual-j_column)
    #
    PreviousFortyEightBitResult = 0
    for i_row in range(BeginInputRow+1,EndRow):                                      # Start at row 9 (one past the label - where the data is), run through 28
        FortyEightBitResult = 0
        for j_column in range(BeginInputColumn,EndInputColumn):                           # Start with column 5 (E).  Run across columns in this row. Repeat out through column 52 (AZ)
            if sheet_input.cell(row=i_row,column=j_column).value == 1:
                FortyEightBitResult = FortyEightBitResult + (2**(EndInputColumnActual-j_column))

            if j_column == EndInputColumnActual:
                # Convert FortyEightBitResult to hex and put in the hex cell
                sheet_input.cell(row=i_row,column=HexColumn).value = "{0:#0{1}x}".format(FortyEightBitResult,14)
                if ( i_row != BeginInputRow+1):
                    if (FortyEightBitResult != PreviousFortyEightBitResult):
                        sheet_input.cell(row=i_row,column=StateChangeColumn).value = "CHANGED"
                        PreviousFortyEightBitResult = FortyEightBitResult

    # Save the spreadsheet (will fail if the spreadsheet is currently open)
    wb.save(file1_name)

    print("\n******* Completed H4RG Timing Plot Generation  *******")
